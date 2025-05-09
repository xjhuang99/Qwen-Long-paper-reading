import os
import re
import time
from datetime import datetime
from typing import List, Dict
import pandas as pd
from openai import OpenAI
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.utils.dataframe import dataframe_to_rows
import concurrent.futures

# Configuration Constants
MAX_INPUT_LENGTH_LONG = 10000000
MAX_REPLY_LENGTH = 6000

COLUMN_NAMES = {
    'A': 'Publication Date',
    'B': 'Author Names',
    'C': 'Journal Name',
    'D': 'Article Title',
    'E': 'Keywords',
    'F': 'Summary',
    'G': 'Core Findings',
    'H': 'Variables',
    'I': 'Theory Name',
    'J': 'Theoretical Framework',
    'K': 'Methodology',
    'L': 'Red Flags',
    'M': 'Relevance to Research'
}


class PDFProcessor:
    @staticmethod
    def process_folder(folder_path: str) -> List[str]:
        """Recursively retrieve all PAF files in a folder and its subfolders"""
        pdf_files = []
        for dirpath, dirnames, filenames in os.walk(folder_path):
            for filename in filenames:
                if filename.lower().endswith('.pdf'):
                    pdf_files.append(os.path.join(dirpath, filename))
        return pdf_files


class QwenAPI:
    def __init__(self, api_key: str, user_prompt_path: str):
        self.client = OpenAI(
            api_key=api_key,
            base_url="https://dashscope.aliyuncs.com/compatible-mode/v1"
        )
        self.user_prompt = self._read_prompt(user_prompt_path)

    def _read_prompt(self, path: str) -> str:
        """Read prompt template from file"""
        with open(path, 'r', encoding='utf-8') as f:
            return f.read()

    def upload_document(self, file_path: str) -> str:
        """Upload file to Qwen API"""
        try:
            with open(file_path, "rb") as f:
                response = self.client.files.create(
                    file=f,
                    purpose="file-extract"
                )
            return response.id
        except Exception as e:
            raise RuntimeError(f"Upload failed: {str(e)}")

    def get_summary(self, file_id: str) -> List[Dict]:
        """Get document summary from Qwen API"""
        max_retries = 3
        for attempt in range(max_retries):
            messages = [
                {'role': 'system', 'content': 'You are a helpful research assistant.'},
                {'role': 'system', 'content': f'fileid://{file_id}'},
                {'role': 'user', 'content': self.user_prompt}
            ]

            try:
                response = self.client.chat.completions.create(
                    model="qwen-long",
                    messages=messages,
                    stream=False
                )
                result = [{"content": response.choices[0].message.content}]
                sections = ResultExporter.parse_sections(result[0]['content'])
                if all(key in sections for key in COLUMN_NAMES):
                    return result
            except Exception as e:
                print(f"API Error (Attempt {attempt + 1}): {str(e)}")

            if attempt < max_retries - 1:
                print(f"Retrying in 5 seconds... (Attempt {attempt + 2})")
                time.sleep(5)

        print("Max retries reached. Returning error response.")
        return [{"error": "API request failed after multiple attempts"}]


class ResultExporter:
    @staticmethod
    def to_excel(data: List[Dict], output_file: str) -> str:  # Modified parameter
        """Export results to Excel with formatting"""
        # Prepare dataframe
        rows = []
        for entry in data:
            for response in entry['raw_responses']:
                row = {
                    'File Path': entry['file_path'],
                    'Process Time': entry['process_time']
                }
                if 'error' in response:
                    row['Error'] = response['error']
                    rows.append(row)
                    continue

                sections = ResultExporter.parse_sections(response['content'])
                for key, col_name in COLUMN_NAMES.items():
                    row[col_name] = ResultExporter.clean_text(sections.get(key, ''))
                rows.append(row)

        df = pd.DataFrame(rows, columns=['File Path', 'Process Time'] + list(COLUMN_NAMES.values()))
        df['Publication Date'] = df['Publication Date'].apply(ResultExporter.format_date)

        # Create formatted workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Research Summary'

        # Add headers
        for col_num, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num, value=col_name)
            cell.font = Font(name='Times New Roman', bold=True)

        # Add data with rich text formatting
        for row_num, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
            for col_num, value in enumerate(row_data, 1):
                if isinstance(value, str):
                    ws.cell(row=row_num, column=col_num).value = ResultExporter.format_rich_text(value)
                else:
                    cell = ws.cell(row=row_num, column=col_num, value=value)
                    cell.font = Font(name='Times New Roman')

        # Apply styling
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = 18

        alignment = Alignment(wrap_text=True, vertical='top')
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = alignment

        ws.row_dimensions[1].height = 30  # Header row
        for row_num in range(2, ws.max_row + 1):
            ws.row_dimensions[row_num].height = 200

        wb.save(output_file)
        print(f"📊 Saved results to: {os.path.abspath(output_file)}")
        return os.path.abspath(output_file)

    @staticmethod
    def format_rich_text(text: str) -> CellRichText:
        """Create rich text formatting for Excel cells"""
        rich_text = CellRichText()
        normal_font = InlineFont(rFont='Times New Roman')
        bold_font = InlineFont(rFont='Times New Roman', b=True)

        patterns = [
            (r'(\d+\.\s*[a-zA-Z\s]+:)', bold_font),
            (r'(Rating for Task \d+)', bold_font)
        ]

        pos = 0
        while pos < len(text):
            matched = False
            for pattern, font in patterns:
                match = re.search(pattern, text[pos:])
                if match:
                    start, end = match.span()
                    if start > 0:
                        rich_text.append(TextBlock(normal_font, text[pos:pos + start]))
                    rich_text.append(TextBlock(font, match.group(1)))
                    pos += end
                    matched = True
                    break
            if not matched:
                rich_text.append(TextBlock(normal_font, text[pos:]))
                break

        return rich_text

    @staticmethod
    def parse_sections(content: str) -> Dict[str, str]:
        """Parse API response into structured sections"""
        sections = {}
        current_key = None
        buffer = []

        for line in content.split('\n'):
            line = line.strip()
            if not line:
                continue

            key_match = re.match(r'^([A-Z])\.\s*(.+)', line)
            if key_match:
                if current_key:
                    sections[current_key] = '\n'.join(buffer).strip()
                current_key = key_match.group(1)
                buffer = [key_match.group(2)]
            else:
                if current_key:
                    buffer.append(line)

        if current_key and buffer:
            sections[current_key] = '\n'.join(buffer).strip()

        return {k: v for k, v in sections.items() if k in COLUMN_NAMES}

    @staticmethod
    def clean_text(text: str) -> str:
        """Clean and normalize text content"""
        text = re.sub(r'###\s*', '', text)
        text = re.sub(r'^\s*[-•]\s*', '', text, flags=re.MULTILINE)
        return re.sub(r'\s+', ' ', text).strip()

    @staticmethod
    def format_date(date_str):
        """Enhanced date formatting with robust type handling"""
        if pd.isna(date_str) or date_str is None:
            return ''
        if isinstance(date_str, (float, int)):
            if isinstance(date_str, float) and date_str.is_integer():
                return str(int(date_str))
            return str(date_str)
        if not isinstance(date_str, str):
            date_str = str(date_str)

        date_str = date_str.strip()
        if not date_str or date_str.lower() == 'nan':
            return ''
        if re.fullmatch(r'\d{4}', date_str):
            return date_str
        try:
            dt = pd.to_datetime(date_str, errors='coerce')
            if pd.notna(dt):
                return dt.strftime('%Y/%m/%d')
        except Exception:
            pass
        return date_str


def process_single_pdf(pdf_path, qwen_client):
    original_filename = os.path.basename(pdf_path)
    try:
        print(f"🔍 Processing: {original_filename}")

        # Check for existing TXT file
        txt_folder = os.path.join(os.path.dirname(pdf_path), "qwen_answers")
        os.makedirs(txt_folder, exist_ok=True)
        txt_path = os.path.join(txt_folder, f"{os.path.splitext(original_filename)[0]}.txt")

        if os.path.exists(txt_path):
            print(f"📁 Using cached response for: {original_filename}")
            with open(txt_path, 'r', encoding='utf-8') as f:
                content = f.read()
        else:
            # Process new file with API
            file_id = qwen_client.upload_document(pdf_path)
            responses = qwen_client.get_summary(file_id)

            if not responses or 'error' in responses[0]:
                raise RuntimeError(responses[0].get('error', 'Unknown API error'))

            content = responses[0]['content']
            # Save API response
            with open(txt_path, 'w', encoding='utf-8') as f:
                f.write(content)

        # Generate new filename from API response
        sections = ResultExporter.parse_sections(content)
        publish_year = ResultExporter.format_date(sections.get('A', '')).split('/')[0] or 'unknown'
        author_names = sections.get('B', '')
        first_author = author_names.split(',')[0].strip() if author_names else ''
        first_author_surname = first_author.split()[-1] if first_author else 'unknown'
        article_title = (sections.get('D', '')[:175] or 'unknown').strip()

        base_name = f"{publish_year}_{first_author_surname}_{article_title}.pdf"
        safe_name = re.sub(r'[\\/*?:"<>|]', '', base_name)
        if not safe_name:
            safe_name = 'unknown.pdf'

        new_pdf_path = os.path.join(os.path.dirname(pdf_path), safe_name)

        os.rename(pdf_path, new_pdf_path)
        print(f"✅ Renamed: {original_filename} -> {os.path.basename(new_pdf_path)}")

        # Update the txt file name to match the new PDF name
        new_txt_name = f"{os.path.splitext(safe_name)[0]}.txt"
        new_txt_path = os.path.join(txt_folder, new_txt_name)
        os.rename(txt_path, new_txt_path)

        return {
            'file_path': new_pdf_path,
            'process_time': datetime.now().strftime("%Y-%m-%d %H:%M"),
            'raw_responses': [{'content': content}]
        }

    except Exception as e:
        print(f"❌ Processing failed for {original_filename}: {str(e)}")
        return {
            'file_path': pdf_path,
            'process_time': datetime.now().strftime("%Y-%m-%d %H:%M"),
            'raw_responses': [{'error': str(e)}]
        }

def main():
    """Main processing workflow"""
    config = {
        "api_key": "xxx",
        "pdf_folder": r"D:\Downloads\PDF\test",
        "user_prompt_path": r"D:\Downloads\PDF\user_prompt.txt",
        "max_workers": 10  # Add parallel processing limit
    }

    processor = PDFProcessor()
    qwen_client = QwenAPI(config["api_key"], config["user_prompt_path"])
    pdf_paths = processor.process_folder(config["pdf_folder"])

    # Generate output filename
    folder_name = os.path.basename(config["pdf_folder"])
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    output_filename = f"research_summary_{folder_name}_{timestamp}.xlsx"
    output_path = os.path.join(config["pdf_folder"], output_filename)

    with concurrent.futures.ThreadPoolExecutor(max_workers=config["max_workers"]) as executor:
        futures = {executor.submit(process_single_pdf, path, qwen_client): path for path in pdf_paths}
        results = [future.result() for future in concurrent.futures.as_completed(futures)]

    ResultExporter.to_excel(results, output_path)
    print(f"\n🎉 Processing complete! Results saved to:\n{output_path}")


if __name__ == "__main__":
    main()
